from .models import Customer
from django.shortcuts import get_object_or_404, render
from openpyxl import load_workbook
from .forms import CustomerForm

# Create your views here.
def index(request):
    return render(request, "customer/index.html")

def customer1(request):
    form = CustomerForm()
    
    ## loading the workbook
    n = 0
    wb = load_workbook("static/docs/EnergyConsumptionDetail(2107).xlsx")
    sheets = wb.sheetnames
    ws = wb[sheets[0]]

    ## customer details and consumption details
    customer_name = ws["B1"].value
    customer_address = ws["B2"].value
    meter_number = ws["B3"].value
    day_rate1 = ws["B6"].value
    day_rate2 = ws["C6"].value
    night_rate1 = ws["B7"].value
    night_rate2 = ws["C7"].value

    ## rate price
    ws = wb[sheets[4]]
    day_rate = ws["B2"].value
    night_rate = ws["B3"].value
    weekend_rate = ws["B4"].value
    weekend_day_rate = ws["B5"].value

    ## energy consumption
    ec_day_rate = day_rate2-day_rate1
    ec_night_rate = night_rate2-night_rate1

    ## consumption cost
    cost_day_rate = ec_day_rate*day_rate
    cost_night_rate = ec_night_rate*night_rate


    ## total energy cost
    total_cost = cost_day_rate+cost_night_rate

    context = {
        "form": form,
        "cust_name": customer_name,
        "cust_add": customer_address,
        "cust_meter": meter_number,
        "cust_day1": day_rate1,
        "cust_day2": day_rate2,
        "cust_night1": night_rate1,
        "cust_night2": night_rate2,
        "day_rate": day_rate,
        "night_rate": night_rate,
        "weekend_rate": weekend_rate,
        "weekend_day_rate": weekend_day_rate,
        "ec_day_rate": ec_day_rate,
        "ec_night_rate": ec_night_rate,
        "cost_day_rate": cost_day_rate,
        "cost_night_rate": cost_night_rate,
        "total_cost": total_cost
    }

    return render(request, "customer/customer1.html", context)

def customer2(request):
    form = CustomerForm()
    ## loading the workbook
    n = 0
    wb = load_workbook("static/docs/EnergyConsumptionDetail(2107).xlsx")
    sheets = wb.sheetnames
    ws = wb[sheets[1]]

    ## customer details and consumption details
    customer_name = ws["B1"].value
    customer_address = ws["B2"].value
    meter_number = ws["B3"].value
    day_rate1 = ws["B6"].value
    day_rate2 = ws["C6"].value
    night_rate1 = ws["B7"].value
    night_rate2 = ws["C7"].value
    weekend_rate1 = ws["B8"].value
    weekend_rate2 = ws["C8"].value

    ## rate price
    ws = wb[sheets[4]]
    day_rate = ws["B2"].value
    night_rate = ws["B3"].value
    weekend_rate = ws["B4"].value
    weekend_day_rate = ws["B5"].value

    ## energy consumption
    ec_day_rate = day_rate2-day_rate1
    ec_night_rate = night_rate2-night_rate1
    ec_weekend_rate = weekend_rate2-weekend_rate1


    ## consumption cost
    cost_day_rate = ec_day_rate*day_rate
    cost_night_rate = ec_night_rate*night_rate
    cost_weekend_rate = ec_weekend_rate*weekend_rate

    ## total energy cost
    total_cost = cost_day_rate+cost_night_rate+cost_weekend_rate
    
    context = {
        "form": form,
        "cust_name": customer_name,
        "cust_add": customer_address,
        "cust_meter": meter_number,
        "cust_day1": day_rate1,
        "cust_day2": day_rate2,
        "cust_night1": night_rate1,
        "cust_night2": night_rate2,
        "cust_weekend1": weekend_rate1,
        "cust_weekend2": weekend_rate2,
        "day_rate": day_rate,
        "night_rate": night_rate,
        "weekend_rate": weekend_rate,
        "weekend_day_rate": weekend_day_rate,
        "ec_day_rate": ec_day_rate,
        "ec_night_rate": ec_night_rate,
        "ec_weekend_rate": ec_weekend_rate,
        "cost_day_rate": cost_day_rate,
        "cost_night_rate": cost_night_rate,
        "cost_weekend_rate": cost_weekend_rate,
        "total_cost": total_cost
    }

    return render(request, "customer/customer2.html", context)

def customer3(request):

    ## loading the workbook
    n = 0
    wb = load_workbook("static/docs/EnergyConsumptionDetail(2107).xlsx")
    sheets = wb.sheetnames
    ws = wb[sheets[2]]

    ## customer details and consumption details
    customer_name = ws["B1"].value
    customer_address = ws["B2"].value
    meter_number = ws["B3"].value
    weekend_day_rate1 = ws["B6"].value
    weekend_day_rate2 = ws["C6"].value
    night_rate1 = ws["B7"].value
    night_rate2 = ws["C7"].value
    weekend_rate1 = ws["B8"].value
    weekend_rate2 = ws["C8"].value

    ## rate price
    ws = wb[sheets[4]]
    day_rate = ws["B2"].value
    night_rate = ws["B3"].value
    weekend_rate = ws["B4"].value
    weekend_day_rate = ws["B5"].value

    ## energy consumption
    ec_weekend_day_rate = weekend_day_rate2-weekend_day_rate1
    ec_night_rate = night_rate2-night_rate1
    ec_weekend_rate = weekend_rate2-weekend_rate1


    ## consumption cost
    cost_weekend_day_rate = ec_weekend_day_rate*weekend_day_rate
    cost_night_rate = ec_night_rate*night_rate
    cost_weekend_rate = ec_weekend_rate*weekend_rate

    ## total energy cost
    total_cost = cost_weekend_day_rate+cost_night_rate+cost_weekend_rate
    
    context = {
        "cust_name": customer_name,
        "cust_add": customer_address,
        "cust_meter": meter_number,
        "cust_weekend_day1": weekend_day_rate1,
        "cust_weekend_day2": weekend_day_rate2,
        "cust_night1": night_rate1,
        "cust_night2": night_rate2,
        "cust_weekend1": weekend_rate1,
        "cust_weekend2": weekend_rate2,
        "day_rate": day_rate,
        "night_rate": night_rate,
        "weekend_rate": weekend_rate,
        "weekend_day_rate": weekend_day_rate,
        "ec_weekend_day_rate": ec_weekend_day_rate,
        "ec_night_rate": ec_night_rate,
        "ec_weekend_rate": ec_weekend_rate,
        "cost_weekend_day_rate": cost_weekend_day_rate,
        "cost_night_rate": cost_night_rate,
        "cost_weekend_rate": cost_weekend_rate,
        "total_cost": total_cost
    }

    return render(request, "customer/customer3.html", context)

def customer4(request):

    ## loading the workbook
    n = 0
    wb = load_workbook("static/docs/EnergyConsumptionDetail(2107).xlsx")
    sheets = wb.sheetnames
    ws = wb[sheets[3]]

    ## customer details and consumption details
    customer_name = ws["B1"].value
    customer_address = ws["B2"].value
    meter_number = ws["B3"].value
    day_rate1 = ws["B6"].value
    day_rate2 = ws["C6"].value
    weekend_rate1 = ws["B7"].value
    weekend_rate2 = ws["C7"].value

    ## rate price
    ws = wb[sheets[4]]
    day_rate = ws["B2"].value
    night_rate = ws["B3"].value
    weekend_rate = ws["B4"].value
    weekend_day_rate = ws["B5"].value

    ## energy consumption
    ec_day_rate = day_rate2-day_rate1
    ec_weekend_rate = weekend_rate2-weekend_rate1


    ## consumption cost
    cost_day_rate = ec_day_rate*day_rate
    cost_weekend_rate = ec_weekend_rate*weekend_rate

    ## total energy cost
    total_cost = cost_day_rate+cost_weekend_rate
    
    context = {
        "cust_name": customer_name,
        "cust_add": customer_address,
        "cust_meter": meter_number,
        "cust_day1": day_rate1,
        "cust_day2": day_rate2,
        "cust_weekend1": weekend_rate1,
        "cust_weekend2": weekend_rate2,
        "day_rate": day_rate,
        "night_rate": night_rate,
        "weekend_rate": weekend_rate,
        "weekend_day_rate": weekend_day_rate,
        "ec_day_rate": ec_day_rate,
        "ec_weekend_rate": ec_weekend_rate,
        "cost_day_rate": cost_day_rate,
        "cost_weekend_rate": cost_weekend_rate,
        "total_cost": total_cost
    }

    return render(request, "customer/customer4.html", context)